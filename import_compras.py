"""
Import script: COMPRAS 2026 spreadsheet → ABMT Commercial System
Imports purchases from COMPRAS sheet into ordens_compra + oc_items.
"""
import sqlite3
import json
import re
import os
from datetime import datetime
from collections import defaultdict
from werkzeug.security import generate_password_hash

import openpyxl

XLSX_PATH = r'C:\Users\pedro\Downloads\COMPRAS 2026 (1).xlsx'
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'comercial.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def classify_product(produto_raw):
    if not produto_raw:
        return 'Diversos', {}
    p = produto_raw.upper().strip()
    specs = {}

    # Transformadores
    if any(k in p for k in ['TRAFO', 'TRANSFORMADOR', 'KVA']):
        if 'NOVO' in p or 'SECO' in p:
            cat = 'Transformador Novo'
        elif 'FUNCIONANDO' in p or 'SEMI' in p or 'NÃO FUNC' in p:
            cat = 'Transformador Usado'
        else:
            cat = 'Transformador Usado'

        pot_match = re.search(r'(\d+)\s*KVA', p)
        if pot_match:
            specs['potencia'] = pot_match.group(1)

        ta_match = re.search(r'(\d+(?:\.\d+)?)\s*KV\b', p)
        if ta_match:
            specs['tensao_alta'] = ta_match.group(1)

        tb_match = re.search(r'(\d+/\d+)', p)
        if tb_match:
            specs['tensao_baixa'] = tb_match.group(1)

        if 'MONO' in p:
            specs['tipo'] = 'Monofásico'
        elif 'TRI' in p:
            specs['tipo'] = 'Trifásico'
        else:
            specs['tipo'] = 'Trifásico'

        if 'SECO' in p:
            specs['isolamento'] = 'A Seco'
        else:
            specs['isolamento'] = 'A Óleo'

        marca_match = re.search(r'(WEG|ABB|SIEMENS|ROMAGNOLE|TRAFO|EATON)', p)
        if marca_match:
            specs['marca'] = marca_match.group(1)

        return cat, specs

    # KVA pattern without TRAFO keyword
    kva_pattern = re.search(r'(\d+)\s*KVA', p)
    if kva_pattern and any(k in p for k in ['KV', 'PEÇA', 'PE\xc7A']):
        specs['potencia'] = kva_pattern.group(1)
        specs['tipo'] = 'Monofásico' if 'MONO' in p else 'Trifásico'
        specs['isolamento'] = 'A Seco' if 'SECO' in p else 'A Óleo'
        if 'FUNCIONANDO' in p:
            return 'Transformador Usado', specs
        return 'Transformador Usado', specs

    if 'BOBINA' in p:
        specs['tipo_aco'] = 'GO' if 'GO' in p else 'GNO' if 'GNO' in p else 'GO'
        larg = re.search(r'(\d+)\s*MM', p)
        if larg:
            specs['largura'] = larg.group(1)
        return 'Bobinas de Aço Silício', specs

    if 'CHAPA' in p or ('SILICIO' in p and 'BOBINA' not in p):
        if 'CORTAD' in p:
            return 'Chapas de Aço Silício Cortadas', specs
        specs['tipo_aco'] = 'GO' if 'GO' in p else 'GNO' if 'GNO' in p else 'GO'
        esp = re.search(r'(\d+[.,]\d+)', p)
        if esp:
            specs['espessura'] = esp.group(1).replace(',', '.')
        return 'Chapas de Aço Silício', specs

    if 'COBRE' in p or 'FIO' in p:
        if 'MEL' in p:
            specs['tipo_cobre'] = 'Mel'
        elif 'MISTO' in p:
            specs['tipo_cobre'] = 'Misto'
        elif 'PAPEL' in p:
            specs['tipo_cobre'] = 'Cobre com Papel'
        return 'Cobre', specs

    if 'ALUMIN' in p or 'INOX' in p:
        return 'Alumínio', specs

    if 'OLEO' in p or 'ÓLEO' in p:
        specs['estado_oleo'] = 'Novo' if 'NOVO' in p else 'Usado'
        return 'Óleo Isolante', specs

    if 'CAIXA' in p or 'NUCLEO' in p or 'NÚCLEO' in p:
        return 'Caixa e Núcleo', specs

    if 'RADIADOR' in p:
        return 'Radiadores', specs

    if 'SUCATA' in p or 'RETALHO' in p:
        return 'Retalho / Sucata', specs

    if 'PAPEL' in p:
        return 'Papel Kraft', specs

    if 'BOMBONA' in p or 'TAMBOR' in p:
        return 'Óleo Isolante', specs

    if 'TERMINAL' in p or 'BUCHA' in p or 'GRAMPO' in p:
        return 'Diversos', specs

    if 'KVA' in p:
        pot = re.search(r'(\d+)\s*KVA', p)
        if pot:
            specs['potencia'] = pot.group(1)
        return 'Transformador Usado', specs

    return 'Diversos', specs


def parse_condicao(cond_raw):
    if not cond_raw:
        return {'tipo': '30/60/90 dias'}
    c = str(cond_raw).strip().upper()
    if 'VISTA' in c or 'PIX' in c or 'A VISTA' in c:
        return {'tipo': 'À vista'}
    if '30/60/90/120' in c:
        return {'tipo': '30/60/90/120 dias'}
    if '30/60/90' in c:
        return {'tipo': '30/60/90 dias'}
    if '30/45/60' in c:
        return {'tipo': '30/45/60 dias'}
    if '28/56' in c:
        return {'tipo': '28/56 dias'}
    if '30/60' in c:
        return {'tipo': '30/60 dias'}
    if '30/45' in c:
        return {'tipo': '30/45/60 dias'}
    if '30 DIAS' in c or '30 DIA' in c:
        return {'tipo': '30 dias'}
    if '60 DIAS' in c or '60 DIA' in c:
        return {'tipo': '30/60 dias'}
    return {'tipo': 'Personalizado', 'texto_livre': str(cond_raw).strip()[:200]}


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    conn = get_db()

    # === 1. Ensure comprador users exist ===
    existing_users = {r['username']: dict(r) for r in conn.execute("SELECT * FROM users").fetchall()}
    user_map = {}

    # Map comprador names → usernames (ALL buyers)
    comprador_username = {
        'PEDRO': 'pedro',
        'PEDRO HENRIQUE': 'pedro_henrique',
        'THIAGO': 'thiago',
        'GUILHERME': 'guilherme',
        'ABMT': 'pedro',       # Compras ABMT = conta do Pedro
        'CIDO': 'cido',
        'CLOVIS': 'clovis',
        'SANDRA': 'sandra',
    }

    for comp_name, username in comprador_username.items():
        if username in existing_users:
            user_map[comp_name] = existing_users[username]['id']
        else:
            nome = comp_name.title()
            conn.execute(
                "INSERT INTO users (username, password_hash, nome, perfil, must_change_password) VALUES (?,?,?,?,1)",
                (username, generate_password_hash('abmt2026'), nome, 'vendedor')
            )
            uid = conn.execute("SELECT last_insert_rowid() as id").fetchone()['id']
            user_map[comp_name] = uid
            existing_users[username] = {'id': uid}
            print(f"  Criado usuário: {username} (id={uid}, nome={nome})")

    conn.commit()
    print(f"\nMapeamento compradores: {user_map}")

    # === 2. Read rows from COMPRAS sheet ===
    ws = wb['COMPRAS']
    all_rows = []

    for r in range(2, ws.max_row + 1):
        cliente = ws.cell(r, 1).value    # CLIENTE (fornecedor)
        data = ws.cell(r, 2).value       # DATA
        oc_num = ws.cell(r, 3).value     # OC
        nota = ws.cell(r, 4).value       # NOTA
        comprador = ws.cell(r, 5).value  # COMPRADOR
        produto = ws.cell(r, 6).value    # PRODUTO
        unidade = ws.cell(r, 7).value    # Un Medida
        qtd = ws.cell(r, 8).value        # QUANTIDADE
        valor_unit = ws.cell(r, 9).value # VALOR (unitário)
        valor_final = ws.cell(r, 10).value # VALOR FINAL
        obs = ws.cell(r, 11).value       # OBS
        forma_pgto = ws.cell(r, 12).value # FORMA DE PAGAMENTO
        comissao = ws.cell(r, 13).value  # COMISSAO

        if not cliente or not comprador or not data:
            continue

        comp_upper = str(comprador).strip().upper()
        if comp_upper not in comprador_username:
            print(f"  WARN: comprador desconhecido '{comp_upper}' na linha {r}, pulando")
            continue

        all_rows.append({
            'row': r,
            'cliente': str(cliente).strip(),
            'data': data,
            'oc_num': str(oc_num).strip() if oc_num else None,
            'nota': str(nota).strip() if nota else None,
            'comprador': comp_upper,
            'produto': str(produto).strip() if produto else '',
            'unidade': str(unidade).strip().upper() if unidade else 'UNIDADE',
            'quantidade': float(qtd) if qtd else 0,
            'valor_unitario': float(valor_unit) if valor_unit and str(valor_unit) not in ('#N/A', 'None', '0') else 0,
            'valor_final': float(valor_final) if valor_final and str(valor_final) not in ('#N/A', 'None', '0') else 0,
            'obs': str(obs).strip() if obs else '',
            'forma_pagamento': str(forma_pgto).strip() if forma_pgto else None,
            'comissao': float(comissao) if comissao and str(comissao) not in ('#N/A', 'None', '0') else 0,
        })

    print(f"\nTotal linhas coletadas: {len(all_rows)}")

    # === 3. Create cadastros for unique fornecedores ===
    existing_cadastros = {}
    for row in conn.execute("SELECT id, razao_social FROM cadastros").fetchall():
        existing_cadastros[row['razao_social'].upper().strip()] = row['id']

    client_names = set(r['cliente'] for r in all_rows)
    cadastro_map = {}
    created_cadastros = 0

    # Get max IMPORT- number
    max_import = conn.execute("SELECT cnpj_cpf FROM cadastros WHERE cnpj_cpf LIKE 'IMPORT-%' ORDER BY cnpj_cpf DESC LIMIT 1").fetchone()
    if max_import:
        import_counter = int(re.search(r'(\d+)', max_import['cnpj_cpf']).group(1)) + 1
    else:
        import_counter = 1

    for name in client_names:
        name_upper = name.upper().strip()
        if name_upper in existing_cadastros:
            cadastro_map[name] = existing_cadastros[name_upper]
            continue

        # Partial match
        found = False
        for existing_name, cid in existing_cadastros.items():
            if name_upper in existing_name or existing_name in name_upper:
                cadastro_map[name] = cid
                found = True
                break
        if found:
            continue

        # Create new cadastro
        conn.execute("""INSERT INTO cadastros (cnpj_cpf, tipo_pessoa, razao_social, endereco_uf, status)
                        VALUES (?, 'PJ', ?, 'SP', 'Ativo')""",
                     (f"IMPORT-{import_counter:04d}", name))
        cid = conn.execute("SELECT last_insert_rowid() as id").fetchone()['id']
        cadastro_map[name] = cid
        existing_cadastros[name_upper] = cid
        created_cadastros += 1
        import_counter += 1

    conn.commit()
    print(f"Cadastros criados: {created_cadastros} (total fornecedores: {len(client_names)})")

    # === 4. Group rows by OC number ===
    oc_groups = defaultdict(list)
    no_oc_counter = 0

    for row in all_rows:
        if row['oc_num'] and row['oc_num'] != 'None':
            key = row['oc_num']
        else:
            no_oc_counter += 1
            key = f"NO_OC_{no_oc_counter}"
        oc_groups[key].append(row)

    print(f"OCs a criar: {len(oc_groups)}")

    # === 5. Create OCs ===
    created_ocs = 0
    created_items = 0
    skipped = 0

    # Get current max OC number
    last_oc = conn.execute("SELECT numero FROM ordens_compra ORDER BY id DESC LIMIT 1").fetchone()
    if last_oc:
        match = re.search(r'(\d+)$', last_oc['numero'])
        oc_counter = int(match.group(1)) + 1 if match else 1
    else:
        oc_counter = 1

    for oc_key, rows in sorted(oc_groups.items()):
        first = rows[0]

        comprador_id = user_map.get(first['comprador'])
        if not comprador_id:
            skipped += len(rows)
            continue

        cadastro_id = cadastro_map.get(first['cliente'])
        if not cadastro_id:
            skipped += len(rows)
            continue

        # Parse date
        if isinstance(first['data'], datetime):
            data_emissao = first['data'].strftime('%Y-%m-%d %H:%M:%S')
        else:
            data_emissao = str(first['data'])

        # Parse payment condition
        condicao = parse_condicao(first['forma_pagamento'])

        # OC number
        if first['oc_num'] and first['oc_num'] != 'None':
            numero = f"OC-{first['oc_num']}"
        else:
            numero = f"OC-{oc_counter:04d}"

        # Check if OC already exists
        exists = conn.execute("SELECT id FROM ordens_compra WHERE numero=?", (numero,)).fetchone()
        if exists:
            skipped += len(rows)
            continue

        # Build obs
        obs_parts = []
        if first['nota']:
            obs_parts.append(f"NF: {first['nota']}")
        for r in rows:
            if r['obs']:
                obs_parts.append(r['obs'])
        observacoes = ' | '.join(obs_parts)[:500] if obs_parts else None

        # Intermediário (comissao)
        total_comissao = sum(r['comissao'] for r in rows)
        interm_comissao_tipo = None
        interm_comissao_valor = None
        if total_comissao > 0:
            interm_comissao_tipo = 'valor_fixo'
            interm_comissao_valor = round(total_comissao, 2)

        # Create OC
        conn.execute("""INSERT INTO ordens_compra (numero, status, cadastro_id, comprador_id,
            numero_omie, data_emissao, condicao_pagamento, forma_pagamento,
            frete, observacoes, intermediario_comissao_tipo, intermediario_comissao_valor)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (numero, 'Recebida Total', cadastro_id, comprador_id,
             first['oc_num'] if first['oc_num'] and first['oc_num'] != 'None' else None,
             data_emissao,
             json.dumps(condicao), 'Faturado',
             'FOB', observacoes,
             interm_comissao_tipo, interm_comissao_valor))

        oc_id = conn.execute("SELECT last_insert_rowid() as id").fetchone()['id']

        # Create items
        for i, row in enumerate(rows):
            categoria, specs = classify_product(row['produto'])

            un = row['unidade']
            if un not in ('KG', 'KVA', 'LITRO', 'UNIDADE', 'TON', 'TONELADA'):
                un = 'UNIDADE'
            if un in ('TON', 'TONELADA'):
                un = 'KG'

            qtd = row['quantidade'] if row['quantidade'] > 0 else 1
            val_unit = row['valor_unitario']
            val_final = row['valor_final']

            if val_final > 0 and val_unit == 0 and qtd > 0:
                val_unit = val_final / qtd

            valor_total = val_final if val_final > 0 else (qtd * val_unit)

            if row['produto']:
                specs['descricao_original'] = row['produto'][:200]

            conn.execute("""INSERT INTO oc_items (oc_id, ordem, categoria, campos_especificos,
                descricao_complementar, quantidade, unidade, valor_unitario, valor_total,
                quantidade_recebida, status)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (oc_id, i, categoria, json.dumps(specs, ensure_ascii=False),
                 row['produto'][:200] if row['produto'] else None,
                 qtd, un, round(val_unit, 2), round(valor_total, 2),
                 qtd, 'Recebido Total'))

            created_items += 1

        created_ocs += 1
        oc_counter += 1

    conn.commit()

    # === 6. Update FTS index ===
    ocs = conn.execute("""SELECT oc.id, oc.numero, c.razao_social, c.cnpj_cpf
                          FROM ordens_compra oc JOIN cadastros c ON oc.cadastro_id=c.id
                          WHERE oc.numero LIKE 'OC-%'""").fetchall()
    for oc in ocs:
        # Check if already indexed
        existing = conn.execute("SELECT 1 FROM busca_global WHERE tipo='oc' AND entidade_id=?", (str(oc['id']),)).fetchone()
        if not existing:
            texto = f"{oc['numero']} {oc['razao_social']} {oc['cnpj_cpf'] or ''}"
            conn.execute("INSERT INTO busca_global(tipo, entidade_id, texto) VALUES (?, ?, ?)",
                         ('oc', str(oc['id']), texto))
    conn.commit()

    # === 7. Summary ===
    print(f"\n{'='*50}")
    print(f"IMPORTAÇÃO DE COMPRAS CONCLUÍDA")
    print(f"{'='*50}")
    print(f"  OCs criadas: {created_ocs}")
    print(f"  Itens criados: {created_items}")
    print(f"  Cadastros criados: {created_cadastros}")
    print(f"  Linhas ignoradas: {skipped}")

    # Per comprador summary
    print(f"\n  Por comprador:")
    for comp_name in sorted(set(r['comprador'] for r in all_rows)):
        vid = user_map.get(comp_name)
        if vid:
            count = conn.execute("SELECT COUNT(*) as c FROM ordens_compra WHERE comprador_id=?", (vid,)).fetchone()['c']
            total = conn.execute("""SELECT COALESCE(SUM(
                (SELECT COALESCE(SUM(valor_total),0) FROM oc_items WHERE oc_id=ordens_compra.id)
            ),0) as t FROM ordens_compra WHERE comprador_id=?""", (vid,)).fetchone()['t']
            print(f"    {comp_name}: {count} OCs, R$ {total:,.2f}")

    # Per month
    print(f"\n  Por mês:")
    monthly = conn.execute("""SELECT strftime('%m', data_emissao) as mes, COUNT(*) as c,
        COALESCE(SUM((SELECT COALESCE(SUM(valor_total),0) FROM oc_items WHERE oc_id=ordens_compra.id)),0) as t
        FROM ordens_compra WHERE status != 'Cancelada' GROUP BY mes ORDER BY mes""").fetchall()
    for m in monthly:
        print(f"    Mês {m['mes']}: {m['c']} OCs, R$ {m['t']:,.2f}")

    conn.close()
    print(f"\nDone!")


if __name__ == '__main__':
    main()
