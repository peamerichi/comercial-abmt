"""
Import script: VENDAS 2026 spreadsheet → ABMT Commercial System
Imports sales for Pedro, Pedro Henrique, Thiago, Adriano from both ABMT and AEB sheets.
"""
import sqlite3
import json
import re
import os
from datetime import datetime
from collections import defaultdict
from werkzeug.security import generate_password_hash

import openpyxl

XLSX_PATH = r'C:\Users\pedro\Downloads\VENDAS 2026  (2).xlsx'
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'comercial.db')
TARGET_VENDEDORES = None  # None = importar TODOS os vendedores

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def parse_uf(uf_raw):
    if not uf_raw:
        return 'SP', False
    uf = str(uf_raw).strip().upper()
    isento = False
    if uf == 'INSERIR' or uf == '':
        return 'SP', False
    # SP18 → SP normal, SP0 → SP isento
    if uf.startswith('SP'):
        if '0' in uf:
            return 'SP', True
        return 'SP', False
    # Remove numbers
    uf_clean = re.sub(r'[^A-Z]', '', uf)
    if len(uf_clean) == 2:
        return uf_clean, isento
    return 'SP', False

def classify_product(produto_raw):
    if not produto_raw:
        return 'Diversos', {}
    p = produto_raw.upper().strip()
    specs = {}

    # Transformadores
    if any(k in p for k in ['TRAFO', 'TRANSFORMADOR', 'KVA']):
        # Detect novo vs usado
        if 'NOVO' in p or 'SECO' in p:
            cat = 'Transformador Novo'
        elif 'FUNCIONANDO' in p or 'SEMI' in p or 'NÃO FUNC' in p:
            cat = 'Transformador Usado'
        else:
            cat = 'Transformador Usado'

        # Extract potencia
        pot_match = re.search(r'(\d+)\s*KVA', p)
        if pot_match:
            specs['potencia'] = pot_match.group(1)

        # Extract tensão alta
        ta_match = re.search(r'(\d+(?:\.\d+)?)\s*KV\b', p)
        if ta_match:
            specs['tensao_alta'] = ta_match.group(1)

        # Extract tensão baixa
        tb_match = re.search(r'(\d+/\d+)', p)
        if tb_match:
            specs['tensao_baixa'] = tb_match.group(1)

        # Tipo
        if 'MONO' in p:
            specs['tipo'] = 'Monofásico'
        elif 'TRI' in p:
            specs['tipo'] = 'Trifásico'
        else:
            specs['tipo'] = 'Trifásico'

        # Isolamento
        if 'SECO' in p:
            specs['isolamento'] = 'A Seco'
        else:
            specs['isolamento'] = 'A Óleo'

        # Marca
        marca_match = re.search(r'(WEG|ABB|SIEMENS|ROMAGNOLE|TRAFO|EATON)', p)
        if marca_match:
            specs['marca'] = marca_match.group(1)

        return cat, specs

    # Also catch patterns like "300KVA - 13.8KV" without TRAFO keyword
    kva_pattern = re.search(r'(\d+)\s*KVA', p)
    if kva_pattern and any(k in p for k in ['KV', 'PEÇA', 'PE\xc7A']):
        specs['potencia'] = kva_pattern.group(1)
        specs['tipo'] = 'Monofásico' if 'MONO' in p else 'Trifásico'
        specs['isolamento'] = 'A Seco' if 'SECO' in p else 'A Óleo'
        if 'FUNCIONANDO' in p:
            return 'Transformador Usado', specs
        return 'Transformador Usado', specs

    # Bobinas
    if 'BOBINA' in p:
        specs['tipo_aco'] = 'GO' if 'GO' in p else 'GNO' if 'GNO' in p else 'GO'
        larg = re.search(r'(\d+)\s*MM', p)
        if larg:
            specs['largura'] = larg.group(1)
        return 'Bobinas de Aço Silício', specs

    # Chapas
    if 'CHAPA' in p or ('SILICIO' in p and 'BOBINA' not in p):
        if 'CORTAD' in p:
            return 'Chapas de Aço Silício Cortadas', specs
        specs['tipo_aco'] = 'GO' if 'GO' in p else 'GNO' if 'GNO' in p else 'GO'
        esp = re.search(r'(\d+[.,]\d+)', p)
        if esp:
            specs['espessura'] = esp.group(1).replace(',', '.')
        return 'Chapas de Aço Silício', specs

    # Cobre
    if 'COBRE' in p or 'FIO' in p:
        if 'MEL' in p:
            specs['tipo_cobre'] = 'Mel'
        elif 'MISTO' in p:
            specs['tipo_cobre'] = 'Misto'
        elif 'PAPEL' in p:
            specs['tipo_cobre'] = 'Cobre com Papel'
        return 'Cobre', specs

    # Alumínio
    if 'ALUMIN' in p or 'INOX' in p:
        return 'Alumínio', specs

    # Óleo
    if 'OLEO' in p or 'ÓLEO' in p:
        specs['estado_oleo'] = 'Novo' if 'NOVO' in p else 'Usado'
        return 'Óleo Isolante', specs

    # Caixa e Núcleo
    if 'CAIXA' in p or 'NUCLEO' in p or 'NÚCLEO' in p:
        return 'Caixa e Núcleo', specs

    # Radiadores
    if 'RADIADOR' in p:
        return 'Radiadores', specs

    # Papel Kraft, Sucata, Retalho
    if 'SUCATA' in p or 'RETALHO' in p:
        return 'Retalho / Sucata', specs

    if 'PAPEL' in p:
        return 'Papel Kraft', specs

    # Bombona/Tambor → Óleo embalagem
    if 'BOMBONA' in p or 'TAMBOR' in p:
        return 'Óleo Isolante', specs

    # Terminal, bucha → Diversos
    if 'TERMINAL' in p or 'BUCHA' in p or 'GRAMPO' in p:
        return 'Diversos', specs

    # Peças de transformador by KVA
    if 'KVA' in p:
        pot = re.search(r'(\d+)\s*KVA', p)
        if pot:
            specs['potencia'] = pot.group(1)
        return 'Transformador Usado', specs

    return 'Diversos', specs


def parse_condicao(cond_raw):
    if not cond_raw:
        return {'tipo': '30/60/90 dias'}
    c = str(cond_raw).strip().upper()
    if 'VISTA' in c or 'PIX' in c or 'A VISTA' in c:
        return {'tipo': 'À vista'}
    if '30/60/90/120' in c:
        return {'tipo': '30/60/90/120 dias'}
    if '30/60/90' in c:
        return {'tipo': '30/60/90 dias'}
    if '30/45/60' in c:
        return {'tipo': '30/45/60 dias'}
    if '28/56' in c:
        return {'tipo': '28/56 dias'}
    if '30/60' in c:
        return {'tipo': '30/60 dias'}
    if '30/45' in c:
        return {'tipo': '30/45/60 dias'}
    if '30 DIAS' in c or '30 DIA' in c:
        return {'tipo': '30 dias'}
    if '60 DIAS' in c or '60 DIA' in c:
        return {'tipo': '30/60 dias'}
    # Personalizado for anything else
    return {'tipo': 'Personalizado', 'texto_livre': str(cond_raw).strip()[:200]}


def parse_frete(transp_raw):
    if not transp_raw:
        return 'FOB', None
    t = str(transp_raw).strip().upper()
    if 'CIF' in t or 'ENTREGA' in t or 'VAMOS' in t:
        return 'CIF', str(transp_raw).strip()
    if 'FOB' in t or 'RETIRA' in t or 'CLIENT' in t:
        return 'FOB', str(transp_raw).strip()
    if 'TRANSPORTADORA' in t:
        return 'FOB', str(transp_raw).strip()
    return 'FOB', str(transp_raw).strip()


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    conn = get_db()

    # === 1. Create missing users ===
    existing_users = {r['username']: dict(r) for r in conn.execute("SELECT * FROM users").fetchall()}
    user_map = {}

    # Map vendedor names to usernames (all sellers)
    vendedor_username = {
        'PEDRO': 'pedro',
        'PEDRO HENRIQUE': 'pedro_henrique',
        'THIAGO': 'thiago',
        'ADRIANO': 'adriano',
        'GUILHERME': 'guilherme',
        'CLOVIS': 'clovis',
        'CLOVIS KAGIL': 'clovis',
        'CLOVIS KAGIL/ GUILHERME': 'guilherme',  # joint sale → Guilherme
        'PRISCILA': 'priscila',
        'RAPHAEL': 'raphael',
        'RAPHAEL PEZZUOL': 'raphael',
    }

    for vend_name, username in vendedor_username.items():
        if username in existing_users:
            user_map[vend_name] = existing_users[username]['id']
        elif username in [v for v in user_map.values()]:
            # Already created in this loop (e.g. 'CLOVIS KAGIL' → 'clovis' already created by 'CLOVIS')
            for k, v in user_map.items():
                if vendedor_username.get(k) == username:
                    user_map[vend_name] = v
                    break
        else:
            nome = vend_name.title()
            conn.execute(
                "INSERT INTO users (username, password_hash, nome, perfil, must_change_password) VALUES (?,?,?,?,1)",
                (username, generate_password_hash('abmt2026'), nome, 'vendedor')
            )
            uid = conn.execute("SELECT last_insert_rowid() as id").fetchone()['id']
            user_map[vend_name] = uid
            existing_users[username] = {'id': uid}
            print(f"  Criado usuário: {username} (id={uid}, nome={nome})")

    conn.commit()
    print(f"\nMapeamento vendedores: {user_map}")

    # === 2. Collect all rows from both sheets ===
    all_rows = []

    for sheet_name, source in [('VENDAS ABMT 2026', 'ABMT'), ('VENDAS AEB 2026', 'AEB')]:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            vendedor = ws.cell(r, 5).value
            cliente = ws.cell(r, 1).value
            data = ws.cell(r, 2).value

            if not vendedor or not cliente or not data:
                continue
            vend_upper = vendedor.strip().upper()
            if vend_upper not in vendedor_username:
                print(f"  WARN: vendedor desconhecido '{vend_upper}' na linha {r} ({sheet_name}), pulando")
                continue
            if 'OV OCUPADA' in str(cliente).upper():
                continue

            ov_num = ws.cell(r, 3).value
            nota = ws.cell(r, 4).value
            produto = ws.cell(r, 6).value
            unidade = ws.cell(r, 7).value
            qtd = ws.cell(r, 8).value
            valor_unit = ws.cell(r, 9).value
            valor_bruto = ws.cell(r, 10).value
            condicao = ws.cell(r, 11).value
            transporte = ws.cell(r, 12).value
            obs = ws.cell(r, 13).value
            uf_raw = ws.cell(r, 14).value

            all_rows.append({
                'source': source,
                'row': r,
                'cliente': str(cliente).strip(),
                'data': data,
                'ov_num': str(ov_num).strip() if ov_num else None,
                'nota': str(nota).strip() if nota else None,
                'vendedor': vend_upper,
                'produto': str(produto).strip() if produto else '',
                'unidade': str(unidade).strip().upper() if unidade else 'UNIDADE',
                'quantidade': float(qtd) if qtd else 0,
                'valor_unitario': float(valor_unit) if valor_unit else 0,
                'valor_bruto': float(valor_bruto) if valor_bruto and str(valor_bruto) not in ('#N/A', 'None', '0') else 0,
                'condicao': condicao,
                'transporte': transporte,
                'obs': str(obs).strip() if obs else '',
                'uf_raw': uf_raw
            })

    print(f"\nTotal linhas coletadas: {len(all_rows)}")

    # === 3. Create cadastros for unique clients ===
    existing_cadastros = {}
    for row in conn.execute("SELECT id, razao_social FROM cadastros").fetchall():
        existing_cadastros[row['razao_social'].upper().strip()] = row['id']

    client_names = set(r['cliente'] for r in all_rows)
    cadastro_map = {}  # client_name → cadastro_id
    created_cadastros = 0

    for name in client_names:
        name_upper = name.upper().strip()
        # Check if exists
        if name_upper in existing_cadastros:
            cadastro_map[name] = existing_cadastros[name_upper]
            continue

        # Also check partial match
        found = False
        for existing_name, cid in existing_cadastros.items():
            if name_upper in existing_name or existing_name in name_upper:
                cadastro_map[name] = cid
                found = True
                break
        if found:
            continue

        # Determine UF from any row for this client
        client_rows = [r for r in all_rows if r['cliente'] == name]
        uf, _ = parse_uf(client_rows[0]['uf_raw']) if client_rows else ('SP', False)

        # Create new cadastro
        conn.execute("""INSERT INTO cadastros (cnpj_cpf, tipo_pessoa, razao_social, endereco_uf, status)
                        VALUES (?, 'PJ', ?, ?, 'Ativo')""",
                     (f"IMPORT-{created_cadastros+1:04d}", name, uf))
        cid = conn.execute("SELECT last_insert_rowid() as id").fetchone()['id']
        cadastro_map[name] = cid
        existing_cadastros[name_upper] = cid
        created_cadastros += 1

    conn.commit()
    print(f"Cadastros criados: {created_cadastros} (total clientes: {len(client_names)})")

    # === 4. Group rows by OV number (source+ov_num = unique key) ===
    ov_groups = defaultdict(list)
    no_ov_counter = 0

    for row in all_rows:
        if row['ov_num'] and row['ov_num'] != 'None':
            key = f"{row['source']}_{row['ov_num']}"
        else:
            no_ov_counter += 1
            key = f"NO_OV_{no_ov_counter}"
        ov_groups[key].append(row)

    print(f"OVs a criar: {len(ov_groups)}")

    # === 5. Create OVs ===
    created_ovs = 0
    created_items = 0
    skipped = 0

    # Get current max OV number
    last_ov = conn.execute("SELECT numero FROM ordens_venda ORDER BY id DESC LIMIT 1").fetchone()
    if last_ov:
        match = re.search(r'(\d+)$', last_ov['numero'])
        ov_counter = int(match.group(1)) + 1 if match else 1
    else:
        ov_counter = 1

    for ov_key, rows in sorted(ov_groups.items()):
        first = rows[0]

        # Get vendedor_id
        vendedor_id = user_map.get(first['vendedor'])
        if not vendedor_id:
            skipped += len(rows)
            continue

        # Get cadastro_id
        cadastro_id = cadastro_map.get(first['cliente'])
        if not cadastro_id:
            skipped += len(rows)
            continue

        # Parse date
        if isinstance(first['data'], datetime):
            data_emissao = first['data'].strftime('%Y-%m-%d %H:%M:%S')
        else:
            data_emissao = str(first['data'])

        # Parse UF
        uf, isento = parse_uf(first['uf_raw'])

        # Parse condition
        condicao = parse_condicao(first['condicao'])

        # Parse frete
        frete, transportadora = parse_frete(first['transporte'])

        # OV number: use original if from ABMT, prefix AEB- for AEB
        if first['source'] == 'ABMT':
            numero = f"OV-{first['ov_num']}" if first['ov_num'] and first['ov_num'] != 'None' else f"OV-{ov_counter:04d}"
        else:
            numero = f"AEB-{first['ov_num']}" if first['ov_num'] and first['ov_num'] != 'None' else f"AEB-{ov_counter:04d}"

        # Check if OV already exists
        exists = conn.execute("SELECT id FROM ordens_venda WHERE numero=?", (numero,)).fetchone()
        if exists:
            skipped += len(rows)
            continue

        # Build obs from all rows
        obs_parts = []
        if first['nota']:
            obs_parts.append(f"NF: {first['nota']}")
        for r in rows:
            if r['obs']:
                obs_parts.append(r['obs'])
        observacoes = ' | '.join(obs_parts)[:500] if obs_parts else None

        # Create OV
        conn.execute("""INSERT INTO ordens_venda (numero, status, cadastro_id, vendedor_id,
            uf_destino, icms_isento, numero_omie, nota_fiscal, data_emissao,
            condicao_pagamento, forma_pagamento, frete, transportadora, observacoes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (numero, 'Aprovada', cadastro_id, vendedor_id,
             uf, 1 if isento else 0,
             first['ov_num'] if first['ov_num'] and first['ov_num'] != 'None' else None,
             first['nota'],
             data_emissao,
             json.dumps(condicao), 'Faturado', frete,
             transportadora, observacoes))

        ov_id = conn.execute("SELECT last_insert_rowid() as id").fetchone()['id']

        # Create items
        for i, row in enumerate(rows):
            categoria, specs = classify_product(row['produto'])

            # Fix unidade
            un = row['unidade']
            if un not in ('KG', 'KVA', 'LITRO', 'UNIDADE'):
                un = 'UNIDADE'

            qtd = row['quantidade'] if row['quantidade'] > 0 else 1
            val_unit = row['valor_unitario']
            val_bruto = row['valor_bruto']

            # If we have bruto but no unit price, calculate it
            if val_bruto > 0 and val_unit == 0 and qtd > 0:
                val_unit = val_bruto / qtd

            valor_total = val_bruto if val_bruto > 0 else (qtd * val_unit)

            # Add preco_unidade for transformadores (unit price)
            if categoria in ('Transformador Usado', 'Transformador Novo') and un == 'KVA':
                potencia = float(specs.get('potencia', 0))
                if potencia > 0 and val_unit > 0:
                    specs['preco_unidade'] = str(round(val_unit * potencia, 2))

            # Store original description in specs
            if row['produto']:
                specs['descricao_original'] = row['produto'][:200]

            conn.execute("""INSERT INTO ov_items (ov_id, ordem, categoria, campos_especificos,
                descricao_complementar, quantidade, unidade, valor_unitario, valor_total, status)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (ov_id, i, categoria, json.dumps(specs, ensure_ascii=False),
                 row['produto'][:200] if row['produto'] else None,
                 qtd, un, round(val_unit, 2), round(valor_total, 2), 'Entregue'))

            created_items += 1

        created_ovs += 1
        ov_counter += 1

    conn.commit()

    # === 6. Update FTS index for new data ===
    ovs = conn.execute("""SELECT ov.id, ov.numero, c.razao_social, c.cnpj_cpf
                          FROM ordens_venda ov JOIN cadastros c ON ov.cadastro_id=c.id
                          WHERE ov.numero LIKE 'OV-%' OR ov.numero LIKE 'AEB-%'""").fetchall()
    for ov in ovs:
        texto = f"{ov['numero']} {ov['razao_social']} {ov['cnpj_cpf'] or ''}"
        # Simple insert (won't duplicate since these are new)
        conn.execute("INSERT INTO busca_global(tipo, entidade_id, texto) VALUES (?, ?, ?)",
                     ('ov', str(ov['id']), texto))
    conn.commit()

    # === 7. Summary ===
    print(f"\n{'='*50}")
    print(f"IMPORTAÇÃO CONCLUÍDA")
    print(f"{'='*50}")
    print(f"  OVs criadas: {created_ovs}")
    print(f"  Itens criados: {created_items}")
    print(f"  Cadastros criados: {created_cadastros}")
    print(f"  Linhas ignoradas: {skipped}")

    # Per vendedor summary
    print(f"\n  Por vendedor:")
    for vend in vendedor_username.keys():
        vid = user_map.get(vend)
        if vid:
            count = conn.execute("SELECT COUNT(*) as c FROM ordens_venda WHERE vendedor_id=?", (vid,)).fetchone()['c']
            total = conn.execute("""SELECT COALESCE(SUM(
                (SELECT COALESCE(SUM(valor_total),0) FROM ov_items WHERE ov_id=ordens_venda.id)
            ),0) as t FROM ordens_venda WHERE vendedor_id=?""", (vid,)).fetchone()['t']
            print(f"    {vend}: {count} OVs, R$ {total:,.2f}")

    # Per month
    print(f"\n  Por mês:")
    monthly = conn.execute("""SELECT strftime('%m', data_emissao) as mes, COUNT(*) as c,
        COALESCE(SUM((SELECT COALESCE(SUM(valor_total),0) FROM ov_items WHERE ov_id=ordens_venda.id)),0) as t
        FROM ordens_venda WHERE status != 'Cancelada' GROUP BY mes ORDER BY mes""").fetchall()
    for m in monthly:
        print(f"    Mês {m['mes']}: {m['c']} OVs, R$ {m['t']:,.2f}")

    conn.close()
    print(f"\nDone!")


if __name__ == '__main__':
    main()
