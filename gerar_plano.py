from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = 'Plano de Ação'

DARK_BG = PatternFill('solid', fgColor='1B1F2B')
HEADER_BG = PatternFill('solid', fgColor='2D3348')
ESSENCIAL_BG = PatternFill('solid', fgColor='1A3A2A')
IMPORTANTE_BG = PatternFill('solid', fgColor='2A2A1A')
AVANCADO_BG = PatternFill('solid', fgColor='1A2A3A')
FUTURO_BG = PatternFill('solid', fgColor='2A1A2A')
SECTION_BG = PatternFill('solid', fgColor='1A73E8')

WHITE = Font(name='Arial', color='FFFFFF', size=10)
WHITE_BOLD = Font(name='Arial', color='FFFFFF', size=10, bold=True)
WHITE_SM = Font(name='Arial', color='FFFFFF', size=9)
HEADER_FONT = Font(name='Arial', color='FFFFFF', size=10, bold=True)
TITLE_FONT = Font(name='Arial', color='FFFFFF', size=16, bold=True)
SUBTITLE_FONT = Font(name='Arial', color='B0B8C8', size=11)
SECTION_FONT = Font(name='Arial', color='FFFFFF', size=11, bold=True)
GREEN_FONT = Font(name='Arial', color='4CAF50', size=10, bold=True)
YELLOW_FONT = Font(name='Arial', color='FFD54F', size=10, bold=True)
BLUE_FONT = Font(name='Arial', color='64B5F6', size=10, bold=True)
PURPLE_FONT = Font(name='Arial', color='CE93D8', size=10, bold=True)
MUTED = Font(name='Arial', color='8892A8', size=9)

thin_border = Border(bottom=Side(style='thin', color='2D3348'))
wrap = Alignment(wrap_text=True, vertical='top')
wrap_center = Alignment(wrap_text=True, vertical='top', horizontal='center')

widths = {'A': 5, 'B': 30, 'C': 34, 'D': 30, 'E': 13, 'F': 13, 'G': 17, 'H': 48, 'I': 34}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.merge_cells('A1:I1')
ws['A1'] = 'PLANO DE AÇÃO — SISTEMA COMERCIAL ABMT'
ws['A1'].font = TITLE_FONT
ws['A1'].fill = DARK_BG
ws['A1'].alignment = Alignment(vertical='center')
ws.row_dimensions[1].height = 36

ws.merge_cells('A2:I2')
ws['A2'] = 'Baseado na Auditoria Executiva de 22/05/2026 — 28 itens priorizados para orientar desenvolvimento'
ws['A2'].font = SUBTITLE_FONT
ws['A2'].fill = DARK_BG
ws.row_dimensions[2].height = 22

ws.merge_cells('A3:I3')
ws['A3'] = 'ESSENCIAL = sem isso não funciona  |  IMPORTANTE = melhora resultado direto  |  AVANÇADO = diferencial competitivo  |  FUTURO = escala enterprise'
ws['A3'].font = MUTED
ws['A3'].fill = DARK_BG
ws.row_dimensions[3].height = 20

headers = ['#', 'Funcionalidade / Melhoria', 'Problema que Resolve', 'Impacto Esperado', 'Prioridade', 'Complexidade', 'Área Afetada', 'Como Deveria Funcionar na Prática', 'Critério para Considerar Pronto']
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_BG
    c.alignment = wrap_center
    c.border = thin_border
ws.row_dimensions[4].height = 32
ws.auto_filter.ref = 'A4:I4'
ws.freeze_panes = 'A5'

items = [
    ('SECTION', 'ESSENCIAL — Implementar nas primeiras 2 semanas'),

    (1, 'Tela "Meu Dia" do Vendedor',
     'Vendedor abre dashboard genérico que não diz o que fazer. Não gera ação diária. Sistema é reativo.',
     'Vendedor abre o sistema todo dia porque vê valor imediato. Adoção sobe de ~30% para 80%+.',
     'Essencial', 'Média', 'Frontend + Backend',
     'Novo endpoint /api/dashboard/meu-dia retorna: follow-ups pendentes (hoje + atrasados), propostas sem ação há 3+ dias, clientes em ciclo de recompra do vendedor, meta vs realizado do mês, parcelas vencendo hoje/amanhã.\n\nFrontend: tela dedicada que substitui o dashboard para perfil vendedor. Cards acionáveis — clicar no cliente abre proposta, clicar no follow-up abre registro.',
     'Vendedor loga e vê pelo menos 3 ações concretas para o dia. Tela carrega em <2s. Cada card leva à ação em 1 clique.'),

    (2, 'Meta por Vendedor com Progresso Visual',
     'Tabela metas existe no banco mas não tem endpoint ativo. Vendedor não sabe quanto falta para bater meta.',
     'Vendedor tem senso de urgência diário. Gerente vê quem está performando sem precisar perguntar.',
     'Essencial', 'Baixa', 'Backend + Frontend',
     'Endpoint /api/metas retorna meta mensal do vendedor, valor realizado (soma OVs do mês), % atingido, projeção linear (realizado / dias passados * dias no mês).\n\nDashboard mostra barra de progresso circular: verde >80%, amarelo 50-80%, vermelho <50%. Meta configurável pelo gestor em Configurações.',
     'Vendedor vê meta e progresso no topo do "Meu Dia". Gestor configura meta por vendedor. Projeção atualiza em tempo real.'),

    (3, 'Motivo de Perda Estruturado',
     'Campo motivo_perda é texto livre. Impossível analisar "por que perdemos" em escala. Dados inutilizáveis para decisão.',
     'Em 30 dias, gestor diz "perdemos 40% por preço". Permite ajustar estratégia comercial com dados reais.',
     'Essencial', 'Baixa', 'Backend + Frontend',
     'Nova coluna motivo_perda_tipo com CHECK(Preço, Prazo de Entrega, Concorrência, Especificação Técnica, Desistência do Cliente, Crédito/Financeiro, Outro).\n\nAo marcar proposta como Perdida: modal com dropdown obrigatório + campo texto complementar + campo "Concorrente" (texto).\n\nEndpoint /api/analytics/perdas retorna distribuição por motivo.',
     'Dropdown é obrigatório ao perder proposta. Relatório de perdas mostra top 3 motivos com valores. Campo concorrente preenchido em >50% das perdas.'),

    (4, 'Campos Obrigatórios no Cadastro',
     'Cadastro aceita apenas CNPJ + razão social. Sem UF, segmento, contato — análises por região/segmento ficam vazias.',
     'Dados limpos desde a entrada. Relatórios por UF e segmento funcionam. Recompra tem whatsapp para contato.',
     'Essencial', 'Baixa', 'Backend + Frontend',
     'Validação no POST/PUT /api/cadastros: obrigatórios = cnpj_cpf, razao_social, endereco_uf, contato_nome, contato_telefone OU contato_whatsapp, segmento.\n\nFrontend: asterisco vermelho nos campos obrigatórios, validação antes do submit com mensagem clara.\n\nCadastros existentes: banner "Cadastro incompleto" com link para editar.',
     'Novo cadastro não salva sem os 5 campos. Cadastros existentes mostram indicador de completude. Em 30 dias, >80% dos cadastros ativos têm segmento.'),

    (5, 'Indicadores de Conversão e Win Rate',
     'Não existe tempo médio de conversão nem taxa de sucesso por vendedor. Gestão é no "achismo".',
     'Gestor identifica gargalos: vendedor rápido vs lento, mês bom vs ruim. Base para coaching individual.',
     'Essencial', 'Baixa', 'Backend',
     'No endpoint /api/pipeline, adicionar:\n- tempo_medio_conversao = AVG(dias entre created_at e data da conversão)\n- win_rate = COUNT(Convertidas) / COUNT(total exceto Rascunho) * 100\n\nAmbos por vendedor e geral. Mostrar no Pipeline Comercial como KPIs no topo da tela.',
     'Pipeline mostra tempo médio em dias e win rate em %. Filtro por vendedor funciona. Dados calculam em <1s.'),

    (6, 'Expiração Automática de Propostas',
     'Propostas com validade vencida ficam abertas para sempre. Pipeline inflado com lixo. Previsão distorcida.',
     'Pipeline reflete realidade. Vendedor recebe alerta antes de perder proposta. Dados de previsão confiáveis.',
     'Essencial', 'Baixa', 'Backend',
     'Job diário (no scheduler existente): propostas com status em (Enviada, Em Negociação) onde data_emissao + validade_dias < hoje → status = Expirada.\n\n2 dias antes: criar notificação "Proposta PROP-XXX expira em 2 dias".\n\nEndpoint /api/propostas/<id>/revalidar para estender validade.',
     'Propostas vencidas mudam automaticamente para Expirada. Notificação 2 dias antes. Revalidação em 1 clique.'),

    ('SECTION', 'IMPORTANTE — Implementar em 30 dias'),

    (7, 'Follow-up Automático por Status',
     'Follow-up é manual e desconectado do funil. Vendedor esquece de cobrar retorno de propostas enviadas.',
     'Nenhuma proposta fica sem acompanhamento. Cadência comercial vira processo, não dependência de memória.',
     'Importante', 'Média', 'Backend + Frontend',
     'Ao mudar status:\n- Enviada → follow-up em 3 dias úteis ("Cobrar retorno de [Cliente]")\n- Em Negociação → follow-up em 5 dias ("Verificar andamento com [Cliente]")\n- Aprovada → follow-up em 1 dia ("Converter proposta [Número]")\n\nFollow-up automático tem flag is_auto=1 e link direto para proposta.',
     'Toda mudança de status gera follow-up com data correta. Aparece no "Meu Dia". Vendedor pode remarcar ou concluir.'),

    (8, 'Pipeline Ponderado (Probabilidade)',
     'Pipeline soma R$ bruto sem considerar chance. Proposta Enviada de R$500k conta igual a Aprovada.',
     'Previsão de faturamento realista. Gestor sabe quanto realmente esperar no mês.',
     'Importante', 'Baixa', 'Backend + Frontend',
     'Probabilidades configuráveis: Rascunho=5%, Enviada=20%, Em Negociação=50%, Aprovada=90%.\n\nPipeline mostra 2 valores: "Pipeline Bruto" (soma total) e "Pipeline Ponderado" (soma × probabilidade).\n\nConfiguração editável em Settings. Previsão usa ponderado + OVs já aprovadas.',
     'Pipeline mostra valor ponderado ao lado do bruto. Diferença >30% entre eles. Config editável pelo gestor.'),

    (9, 'DRE Comercial Simplificado',
     'Não existe visão de resultado: receita − custo − comissão − impostos. Diretor não sabe resultado real.',
     'Diretor vê resultado comercial por mês sem exportar para Excel. Decisões de preço baseadas em dados.',
     'Importante', 'Média', 'Backend + Frontend',
     'Endpoint /api/analytics/dre?mes=X&ano=Y calcula:\n- Receita Bruta (soma OV items)\n- (-) PIS/COFINS (% configurado)\n- (-) ICMS (por UF)\n- = Receita Líquida\n- (-) Custo Mercadoria (campo custo dos OV items)\n- = Margem Bruta\n- (-) Comissões\n- = Resultado Operacional\n\nTabela 12 meses + gráfico evolução.',
     'DRE gera em <3s. Valores batem com fechamento existente. 12 meses lado a lado. Margem % por categoria visível.'),

    (10, 'Margem Consolidada no Dashboard',
     'Margem está no item individual. Não existe visão consolidada por categoria/vendedor/mês.',
     'Gestor detecta erosão de margem antes que vire prejuízo. Vendedor que desconta demais fica visível.',
     'Importante', 'Baixa', 'Backend + Frontend',
     'No dashboard avançado, seção "Margem":\nmargem_media = SUM(margem × valor_total) / SUM(valor_total) por categoria e vendedor.\n\nAlerta se margem <15% em qualquer categoria.\nCard no dashboard gestor com margem do mês vs mês anterior.',
     'Dashboard gestor mostra margem por categoria. Alerta visual abaixo de 15%. Comparativo mês anterior.'),

    (11, 'Aging de Contas a Receber',
     'Parcelas mostram vencidas vs pendentes, sem aging (1-30, 31-60, 61-90, 90+ dias).',
     'Gestor prioriza cobrança por urgência. Identifica clientes crônicos vs atrasos pontuais.',
     'Importante', 'Baixa', 'Backend + Frontend',
     'Endpoint /api/parcelas/aging retorna buckets: Em dia, 1-30 vencido, 31-60, 61-90, 90+.\nCada bucket: total R$, qtd parcelas, lista clientes.\n\nGráfico barras empilhadas + tabela. Drill-down: clicar bucket mostra parcelas.',
     'Aging carrega em <2s. Buckets corretos (validar 3 parcelas manual). Drill-down funciona. Total bate.'),

    (12, 'Renomear "IA Insights" → "Consulta Rápida"',
     'Nome "IA" cria expectativa que não entrega. É keyword matching — funciona bem mas nome engana.',
     'Expectativa alinhada com realidade. Usuário confia na ferramenta porque entrega o que promete.',
     'Importante', 'Baixa', 'Frontend',
     'Renomear em todo frontend: sidebar, assistente panel, tooltips.\nMudar ícone de "brain" para "search" ou "zap".\nManter lógica backend intacta.\nAtualizar sugestões rápidas.',
     'Zero referência a "IA" no sistema. Nome "Consulta Rápida" em todas as telas. Funcionalidade idêntica.'),

    (13, 'Alerta de Margem Baixa na Conversão',
     'OV criada com margem <10% e ninguém percebe até o fechamento. Prejuízo silencioso.',
     'Gestor revisa antes de faturar. Vendedor pensa duas vezes antes de dar desconto absurdo.',
     'Importante', 'Baixa', 'Backend + Frontend',
     'Na conversão (POST /converter): calcular margem de cada item.\nSe margem <15%: retornar warning (não bloquear).\nFrontend: banner amarelo "Item X tem margem de 8%".\n\nSe perfil=vendedor e margem<10%: bloquear, exigir aprovação do gestor.',
     'Conversão com margem baixa mostra alerta. Vendedor não converte <10% sem aprovação. Log registra alertas.'),

    (14, 'Concentração de Carteira',
     'Se 60% do faturamento vem de 2 clientes e 1 sai, empresa quebra. Ninguém monitora.',
     'Risco de dependência visível. Força diversificação proativa da carteira.',
     'Importante', 'Baixa', 'Backend + Frontend',
     'No analytics, calcular: % faturamento dos top 3 clientes.\nSe >40% em 1 cliente ou >60% nos top 3 → alerta vermelho.\n\nCard no dashboard diretor: "Top 3 = X% do faturamento" com indicador risco.',
     'Indicador mostra % correto. Alerta quando >40%/60%. Atualiza mensalmente.'),

    (15, 'Consolidar Dashboards em Tabs',
     '3 telas de dashboard separadas confundem. Usuário não sabe onde achar o que precisa.',
     'Navegação simplificada. 1 lugar para tudo de visão geral.',
     'Importante', 'Média', 'Frontend',
     'Dashboard vira tela única com 3 tabs:\n- "Resumo" (dashboard atual)\n- "Análise" (avançado com filtros)\n- "Comparativo" (mês/trimestre/ano)\n\nTabs no topo, troca sem reload. Mesmos endpoints, só reorganizar frontend.',
     'Tudo em 1 tela. Tabs trocam instantaneamente (cache). Nenhuma funcionalidade perdida.'),

    ('SECTION', 'AVANÇADO — Implementar em 60 dias'),

    (16, 'Migração SQLite → PostgreSQL',
     'SQLite tem write lock global. 5+ usuários simultâneos causam erros e lentidão em pico.',
     'Sistema aguenta 20+ usuários simultâneos. Preparado para crescimento.',
     'Avançado', 'Alta', 'Backend + Infra',
     'Substituir sqlite3 por psycopg2 ou SQLAlchemy.\nAdaptar queries strftime() → extract()/to_char().\nScript de migração de dados.\nDocker compose com PostgreSQL.',
     'Todos endpoints idênticos. 10 writes simultâneos sem erro. Migration sem perda de dados.'),

    (17, 'Separar app.py em Blueprints',
     'Monolito de 5300 linhas. Bug em fechamento pode derrubar login. Manutenção insustentável.',
     'Desenvolvimento paralelo possível. Bugs isolados. Code review viável.',
     'Avançado', 'Média', 'Backend',
     'Blueprints: auth.py, cadastros.py, propostas.py, ordens.py, financeiro.py, analytics.py, config.py.\nCada um registrado no app.py principal (~50 linhas).\nURLs mantidas idênticas.',
     'app.py tem <100 linhas. Cada blueprint <800 linhas. URLs idênticas. Zero mudança no frontend.'),

    (18, 'Testes Automatizados (Críticos)',
     'Zero testes. Qualquer mudança pode quebrar comissão, parcelas ou conversão sem detecção.',
     'Confiança para evoluir rápido. Deploy sem medo. Regressão detectada automaticamente.',
     'Avançado', 'Média', 'Backend',
     'Pytest + Flask test_client. Mínimo:\n1. Conversão proposta→OV gera parcelas corretas\n2. Comissão por perfil/categoria\n3. Baixa parcial atualiza status\n4. Pipeline retorna dados consistentes\n5. Meta vs realizado bate com OVs\n\nCI via GitHub Actions a cada push.',
     '15+ testes cobrindo fluxos críticos. Todos passam. CI configurado. Cobertura >60% em conversão e financeiro.'),

    (19, 'Push Notifications Reais',
     'Notificações só aparecem quando abre o app. Não busca o vendedor proativamente.',
     'Vendedor recebe alerta no celular sem abrir app. Follow-up, proposta, parcela.',
     'Avançado', 'Alta', 'Backend + Frontend + Infra',
     'Web Push API + VAPID keys.\n/api/push/subscribe salva subscription.\nAo criar notificação, enviar push via pywebpush.\n\nEventos: follow-up vencendo, proposta expira, parcela vence, proposta aprovada.',
     'Push funciona em Android Chrome e iOS Safari. Opt-in funciona. 4+ tipos de evento geram push.'),

    (20, 'Relatório Ad-Hoc (Explorador)',
     'Relatórios pré-formatados. Não responde "vendas de Cobre no PR nos últimos 6 meses".',
     'Gestor responde suas perguntas sem pedir ao desenvolvedor. Autonomia analítica.',
     'Avançado', 'Alta', 'Backend + Frontend',
     'Tela "Explorador de Dados" com filtros combináveis: período, categoria (multi), vendedor, UF, cliente, tipo.\n\n/api/analytics/explorer aceita filtros via query params. Retorna total, quantidade, por mês, top itens. Export CSV.',
     '5+ combinações de filtro funcionam. Resultado em <3s. Export CSV OK. Filtros persistem.'),

    (21, 'Registro de Concorrente',
     'Perde proposta sem registrar para quem perdeu. Inteligência competitiva zero.',
     'Em 90 dias, sabe quais concorrentes mais tiram negócios e em quais categorias.',
     'Avançado', 'Baixa', 'Backend + Frontend',
     'Campo concorrente_nome na tabela propostas. Aparece no modal de perda com motivo.\n\n/api/analytics/concorrentes: ranking por frequência e valor perdido, cruzado com categoria e região.',
     'Campo existe no modal de perda. Relatório gera dados após 30+ propostas perdidas com preenchimento.'),

    (22, 'Churn Rate e Saúde da Base',
     'Lista de inativos 90d existe, mas sem taxa de churn nem tendência.',
     'Métrica clara: "perdemos 3, ganhamos 5 = base cresceu". Tendência visível.',
     'Avançado', 'Média', 'Backend + Frontend',
     'Cliente ativo = comprou nos últimos 90 dias.\n/api/analytics/churn: ativos início mês, novos, churned, reativados.\nTaxa = churned / ativos início.\nGráfico 12 meses.',
     'Churn correto (validar 2 meses manual). Gráfico 12 meses. Novos vs churned visíveis.'),

    ('SECTION', 'FUTURO — Implementar em 90+ dias'),

    (23, 'Integração Omie (NF + Parcelas)',
     'NF é campo texto manual. Parcelas em 2 sistemas. Retrabalho e erro humano.',
     'Elimina retrabalho. NF sincroniza. Parcela paga no Omie reflete no sistema.',
     'Futuro', 'Alta', 'Backend + Infra',
     'API Omie: ao faturar OV, buscar NF e preencher auto.\nSincronizar parcelas: paga no Omie → atualiza status.\nPolling 30min.\nOV.numero_omie = chave de ligação.',
     'NF aparece auto na OV. Parcela paga no Omie muda em <30min. Zero entrada manual.'),

    (24, 'Multi-Empresa (ABMT + AEB)',
     'Sistema serve só ABMT. AEB sem sistema ou usa planilha.',
     'Uma plataforma, duas empresas. Dados isolados, visão consolidada para diretoria.',
     'Futuro', 'Alta', 'Backend + Frontend',
     'Coluna empresa_id em tabelas principais.\nTenant middleware por request.\nSeleção empresa no login ou switch topbar.\nDashboard consolidado para diretor.',
     'Login seleciona empresa. Dados isolados. Diretor vê consolidado. Zero vazamento entre empresas.'),

    (25, 'API Documentada (Swagger)',
     'Sem documentação de API. Integrações dependem de ler código fonte.',
     'Qualquer dev ou ferramenta consome a API sem perguntar. Integração BI.',
     'Futuro', 'Média', 'Backend',
     'Flask-RESTX ou Flasgger gera Swagger auto.\nTodos endpoints documentados: URL, método, params, response, auth.\nServir em /api/docs.',
     'Swagger UI em /api/docs. Todos endpoints documentados. Exemplos request/response.'),

    (26, 'Relatório Semanal por Email',
     'Gestor precisa abrir sistema para ver resultados. Nenhum resumo proativo.',
     'Toda segunda, gestor recebe resumo sem abrir nada. Mantém pulso na operação.',
     'Futuro', 'Média', 'Backend + Infra',
     'Scheduler segunda 7h: gerar HTML com faturamento semana, pipeline, inadimplência, clientes risco, top vendedor.\nEnviar SMTP (configurável).\nOpt-in por usuário.',
     'Email chega toda segunda. HTML OK em Gmail/Outlook. Opt-in funciona. Dados batem com dashboard.'),

    (27, 'Categorias Configuráveis',
     'CATEGORIAS é array hardcoded no Python. Mudar mix exige deploy.',
     'Gestor adiciona/edita categorias sem desenvolvedor. Sistema adapta ao negócio.',
     'Futuro', 'Média', 'Backend + Frontend',
     'Migrar CATEGORIAS para tabela configuracoes (JSON).\n/api/config/categorias para CRUD.\nTela editável em Configurações.\nCompatibilidade com dados históricos.',
     'Gestor adiciona categoria e aparece em propostas/OVs. Dados antigos OK. Sem deploy.'),

    (28, 'Analytics de Uso Interno',
     'Não sabe se vendedores usam o sistema. Pode investir em features que ninguém acessa.',
     'Decisão de produto baseada em dados reais. Features abandonadas visíveis.',
     'Futuro', 'Média', 'Backend',
     'Middleware registra usage_log: user_id, endpoint, timestamp.\n/api/admin/usage: logins/dia, telas top, horários pico, features sem uso.\nDashboard admin.',
     'Log ativo sem impacto (<5ms). Dashboard mostra 7+ dias. Top 10 e bottom 5 visíveis.'),
]

row = 5
for item in items:
    if item[0] == 'SECTION':
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'] = item[1]
        ws[f'A{row}'].font = SECTION_FONT
        ws[f'A{row}'].fill = SECTION_BG
        ws[f'A{row}'].alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 28
        row += 1
        continue

    id_num, nome, problema, impacto, prioridade, complexidade, area, como, criterio = item

    if prioridade == 'Essencial':
        bg = ESSENCIAL_BG
        prio_font = GREEN_FONT
    elif prioridade == 'Importante':
        bg = IMPORTANTE_BG
        prio_font = YELLOW_FONT
    elif prioridade == 'Avançado':
        bg = AVANCADO_BG
        prio_font = BLUE_FONT
    else:
        bg = FUTURO_BG
        prio_font = PURPLE_FONT

    if complexidade == 'Baixa':
        comp_font = Font(name='Arial', color='4CAF50', size=10)
    elif complexidade == 'Média':
        comp_font = Font(name='Arial', color='FFD54F', size=10)
    else:
        comp_font = Font(name='Arial', color='EF5350', size=10)

    vals = [id_num, nome, problema, impacto, prioridade, complexidade, area, como, criterio]
    for col_idx, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.fill = bg
        c.alignment = wrap if col_idx > 1 else wrap_center
        c.border = thin_border
        if col_idx == 1:
            c.font = WHITE_BOLD
        elif col_idx == 2:
            c.font = WHITE_BOLD
        elif col_idx == 5:
            c.font = prio_font
        elif col_idx == 6:
            c.font = comp_font
        else:
            c.font = WHITE_SM

    ws.row_dimensions[row].height = 95
    row += 1

# === RESUMO EXECUTIVO ===
ws2 = wb.create_sheet('Resumo Executivo')
ws2.sheet_properties.tabColor = '1A73E8'

for col, w in {'A': 22, 'B': 16, 'C': 18, 'D': 18, 'E': 18}.items():
    ws2.column_dimensions[col].width = w

ws2.merge_cells('A1:E1')
ws2['A1'] = 'RESUMO EXECUTIVO — PLANO DE AÇÃO'
ws2['A1'].font = TITLE_FONT
ws2['A1'].fill = DARK_BG
ws2.row_dimensions[1].height = 36

ws2.merge_cells('A2:E2')
ws2['A2'] = ''
ws2['A2'].fill = DARK_BG

summary_headers = ['Classificação', 'Qtd Itens', 'Baixa', 'Média', 'Alta']
for col_idx, h in enumerate(summary_headers, 1):
    c = ws2.cell(row=3, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_BG
    c.alignment = Alignment(horizontal='center')

summary_data = [
    ('Essencial', 6, 4, 2, 0, ESSENCIAL_BG, GREEN_FONT),
    ('Importante', 9, 6, 2, 1, IMPORTANTE_BG, YELLOW_FONT),
    ('Avançado', 7, 1, 2, 4, AVANCADO_BG, BLUE_FONT),
    ('Futuro', 6, 0, 4, 2, FUTURO_BG, PURPLE_FONT),
]

for i, (cat, qtd, b, m, a, bg, f) in enumerate(summary_data):
    r = 4 + i
    for col_idx, val in enumerate([cat, qtd, b, m, a], 1):
        c = ws2.cell(row=r, column=col_idx, value=val)
        c.fill = bg
        c.font = f if col_idx == 1 else WHITE
        c.alignment = Alignment(horizontal='center')

r = 8
c = ws2.cell(row=r, column=1, value='TOTAL')
c.font = WHITE_BOLD
c.fill = HEADER_BG
c.alignment = Alignment(horizontal='center')
for col_idx, val in enumerate([28, 11, 10, 7], 2):
    c = ws2.cell(row=r, column=col_idx, value=val)
    c.font = WHITE_BOLD
    c.fill = HEADER_BG
    c.alignment = Alignment(horizontal='center')

ws2.merge_cells('A10:E10')
ws2['A10'] = 'CRONOGRAMA DE IMPLEMENTAÇÃO'
ws2['A10'].font = Font(name='Arial', color='FFFFFF', size=14, bold=True)
ws2['A10'].fill = DARK_BG
ws2.row_dimensions[10].height = 32

timeline_headers = ['Fase', 'Período', 'Itens', 'Foco', 'Métrica de Sucesso']
for col_idx, h in enumerate(timeline_headers, 1):
    c = ws2.cell(row=11, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_BG
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

timeline = [
    ('Fase 1', 'Semana 1-2', '#1 a #6', 'Adoção do vendedor', 'Vendedor abre app 1x/dia', ESSENCIAL_BG),
    ('Fase 2', 'Semana 3-4', '#7 a #15', 'Gestão e visibilidade', 'Pipeline ponderado ativo', IMPORTANTE_BG),
    ('Fase 3', 'Semana 5-8', '#16 a #22', 'Escala e confiança', '10 writes simultâneos OK', AVANCADO_BG),
    ('Fase 4', 'Semana 9-12', '#23 a #28', 'Enterprise', 'Integração Omie ativa', FUTURO_BG),
]

for i, (fase, periodo, itens, foco, metrica, bg) in enumerate(timeline):
    r = 12 + i
    for col_idx, val in enumerate([fase, periodo, itens, foco, metrica], 1):
        c = ws2.cell(row=r, column=col_idx, value=val)
        c.fill = bg
        c.font = WHITE
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws2.row_dimensions[r].height = 30

# Key insight
ws2.merge_cells('A17:E17')
ws2['A17'] = 'MENSAGEM-CHAVE: A prioridade #1 não é técnica — é ADOÇÃO. Se vendedor não abre todo dia, nada mais importa.'
ws2['A17'].font = Font(name='Arial', color='FFD54F', size=11, bold=True)
ws2['A17'].fill = DARK_BG
ws2.row_dimensions[17].height = 30

out = 'Plano_Acao_ABMT_Comercial.xlsx'
wb.save(out)
print(f'OK: {out}')
